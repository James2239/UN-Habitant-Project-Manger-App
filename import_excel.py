import openpyxl
import sqlite3
from datetime import datetime

DB_FILE = 'projects.db'
EXCEL_FILE = 'Application Development - Exam Data.xlsx'

# Helper to get or insert and return ID
def get_or_create(cursor, table, column, value):
    cursor.execute(f"SELECT {table}ID FROM {table} WHERE {column} = ?", (value,))
    row = cursor.fetchone()
    if row:
        return row[0]
    cursor.execute(f"INSERT INTO {table} ({column}) VALUES (?)", (value,))
    return cursor.lastrowid

def main():
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active
    conn = sqlite3.connect(DB_FILE)
    cur = conn.cursor()
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    print('Headers:', headers)
    row_count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        data = dict(zip(headers, row))
        print('Importing row:', data)
        # Insert Project
        try:
            cur.execute('''INSERT OR IGNORE INTO Project (ProjectID, ProjectTitle, PAASCode, ApprovalStatus, Fund, PAGValue, StartDate, EndDate, LeadOrgUnit, TotalExpenditure, TotalContribution, TotalContributionMinusExpenditure, TotalPSC)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)''', (
                data['ProjectID'], data['Project Title'], data['PAAS Code'], data['Approval Status'], data['Fund'], data['PAG Value'],
                data['Start Date'].strftime('%Y-%m-%d') if data['Start Date'] else None,
                data['End Date'].strftime('%Y-%m-%d') if data['End Date'] else None,
                data['Lead Org Unit'], data['Total Expenditure'], data['Total Contribution'],
                data['Total Contribution - Total Expenditure'], data['Total PSC']
            ))
        except Exception as e:
            print(f'Error inserting project row: {e}')
            continue
        # Countries (can be comma or semicolon separated)
        try:
            if data['Country(ies)']:
                for country in [c.strip() for c in str(data['Country(ies)']).replace(';', ',').split(',') if c.strip()]:
                    country_id = get_or_create(cur, 'Country', 'CountryName', country)
                    cur.execute('INSERT OR IGNORE INTO ProjectCountry (ProjectID, CountryID) VALUES (?, ?)', (data['ProjectID'], country_id))
        except Exception as e:
            print(f'Error inserting country for row {data}: {e}')
        # Themes
        try:
            if data['Theme(s)']:
                for theme in [t.strip() for t in str(data['Theme(s)']).replace(';', ',').split(',') if t.strip()]:
                    theme_id = get_or_create(cur, 'Theme', 'ThemeName', theme)
                    cur.execute('INSERT OR IGNORE INTO ProjectTheme (ProjectID, ThemeID) VALUES (?, ?)', (data['ProjectID'], theme_id))
        except Exception as e:
            print(f'Error inserting theme for row {data}: {e}')
        # Donors
        try:
            if data['Donor(s)']:
                for donor in [d.strip() for d in str(data['Donor(s)']).replace(';', ',').split(',') if d.strip()]:
                    donor_id = get_or_create(cur, 'Donor', 'DonorName', donor)
                    cur.execute('INSERT OR IGNORE INTO ProjectDonor (ProjectID, DonorID) VALUES (?, ?)', (data['ProjectID'], donor_id))
        except Exception as e:
            print(f'Error inserting donor for row {data}: {e}')
        row_count += 1
    conn.commit()
    conn.close()
    print(f'Import complete. {row_count} rows imported.')

if __name__ == '__main__':
    main()
